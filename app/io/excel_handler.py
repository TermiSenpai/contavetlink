"""Lectura y escritura Excel (.xlsx) usando openpyxl.

Cobertura v1.0:
  - Import/Export de mappings_clientes
  - Import/Export de mappings_articulos
  - Export de previews de exportación (Fase 4)
  - Export del historial

Las cabeceras y la semántica son **idénticas** a las del CSV (mismas columnas,
mismo orden, misma validación). El round-trip CSV ↔ Excel es trivial.
"""
from __future__ import annotations

import logging
import sqlite3
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.io.csv_handler import (
    ARTICULO_FIELDS,
    CLIENTE_FIELDS,
    HISTORIAL_FIELDS,
    ResultadoImport,
    _importar_fila_articulo,
    _importar_fila_cliente,
    _serializar,
)
from app.mapping.store import SubcuentaInvalidaError

log = logging.getLogger(__name__)


# ─── Mappings de clientes ──────────────────────────────────────────────────


def importar_mappings_clientes(conn: sqlite3.Connection, ruta: Path) -> ResultadoImport:
    return _importar_xlsx(
        conn, ruta, CLIENTE_FIELDS, _importar_fila_cliente,
    )


def exportar_mappings_clientes(conn: sqlite3.Connection, ruta: Path) -> Path:
    rows = conn.execute(
        f"SELECT {', '.join(CLIENTE_FIELDS)} FROM mappings_clientes "
        "ORDER BY codigo_gesdai"
    ).fetchall()
    _escribir_xlsx(ruta, 'clientes', CLIENTE_FIELDS, rows)
    return ruta


# ─── Mappings de artículos ─────────────────────────────────────────────────


def importar_mappings_articulos(conn: sqlite3.Connection, ruta: Path) -> ResultadoImport:
    return _importar_xlsx(
        conn, ruta, ARTICULO_FIELDS, _importar_fila_articulo,
    )


def exportar_mappings_articulos(conn: sqlite3.Connection, ruta: Path) -> Path:
    rows = conn.execute(
        f"SELECT {', '.join(ARTICULO_FIELDS)} FROM mappings_articulos "
        "ORDER BY clave_gesdai"
    ).fetchall()
    _escribir_xlsx(ruta, 'articulos', ARTICULO_FIELDS, rows)
    return ruta


# ─── Historial ─────────────────────────────────────────────────────────────


def exportar_historial(conn: sqlite3.Connection, ruta: Path) -> Path:
    rows = conn.execute(
        f"SELECT {', '.join(HISTORIAL_FIELDS)} FROM exportaciones "
        "ORDER BY fecha_export DESC"
    ).fetchall()
    _escribir_xlsx(ruta, 'historial', HISTORIAL_FIELDS, rows)
    return ruta


def exportar_preview(facturas_resueltas: list[dict], ruta: Path) -> Path:
    """Exporta una vista previa de exportación con su semáforo a Excel."""
    raise NotImplementedError("Pendiente de implementar en Fase 4")


# ─── Helpers internos ──────────────────────────────────────────────────────


def _importar_xlsx(
    conn: sqlite3.Connection,
    ruta: Path,
    columnas_esperadas: tuple[str, ...],
    importador_fila,
) -> ResultadoImport:
    """Loop común de import. Reusa los importadores fila-a-fila del CSV.

    Esto garantiza que CSV y Excel comparten exactamente la misma semántica
    (validación, no sobreescritura de revisados, errores acumulados).
    """
    wb = load_workbook(ruta, read_only=True, data_only=True)
    try:
        sheet = wb.active
        filas = sheet.iter_rows(values_only=True)
        cabecera = next(filas, None)
        if cabecera is None:
            raise ValueError(f"{ruta.name}: hoja vacía")
        cabecera = [str(c).strip() if c is not None else '' for c in cabecera]
        faltan = [c for c in columnas_esperadas if c not in cabecera]
        if faltan:
            raise ValueError(
                f"{ruta.name}: faltan columnas {', '.join(faltan)}"
            )
        idx = {c: cabecera.index(c) for c in columnas_esperadas}

        resultado = ResultadoImport()
        for nlinea, fila in enumerate(filas, start=2):
            if fila is None or all(v is None or v == '' for v in fila):
                continue
            try:
                fila_dict = {c: _celda_a_str(fila[idx[c]]) for c in columnas_esperadas}
                importador_fila(conn, fila_dict, resultado)
            except (SubcuentaInvalidaError, ValueError) as e:
                resultado.errores.append(f"línea {nlinea}: {e}")
        conn.commit()
        return resultado
    finally:
        wb.close()


def _escribir_xlsx(
    ruta: Path,
    sheet_name: str,
    columnas: tuple[str, ...],
    rows: list,
) -> None:
    ruta.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    sheet = wb.active
    sheet.title = sheet_name
    sheet.append(list(columnas))
    for r in rows:
        sheet.append([_serializar(r[col]) for col in columnas])
    wb.save(ruta)


def _celda_a_str(valor: object) -> str:
    """Normaliza una celda de openpyxl a string. None y vacío → ''.

    Los enteros/floats se convierten con `str()` — el round-trip CSV/Excel
    no debería perder precisión porque los campos del intermediario son
    todos texto/booleano/entero pequeño.
    """
    if valor is None:
        return ''
    if isinstance(valor, bool):
        return '1' if valor else '0'
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor)


__all__ = [
    'importar_mappings_clientes',
    'exportar_mappings_clientes',
    'importar_mappings_articulos',
    'exportar_mappings_articulos',
    'exportar_historial',
    'exportar_preview',
]
