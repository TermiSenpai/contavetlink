"""Tests Fase 2 — store + sync + import/export CSV y Excel.

El plan mete aquí pruebas que técnicamente cubren store/sync además del IO
en sí, pero todas comparten el mismo `db_conn` fixture y testean la misma
superficie (mappings + sincronización idempotente), así que viven juntas.
"""
from __future__ import annotations

from pathlib import Path

import pytest

from app.io import csv_handler, excel_handler
from app.mapping import sync as sync_mod
from app.mapping.store import (
    SubcuentaInvalidaError,
    get_articulo_mapping,
    get_cliente_mapping,
    list_clientes_mappings,
    marcar_articulo_revisado,
    marcar_cliente_revisado,
    progreso_clientes,
    set_cuenta_articulo,
    set_subcuenta_cliente,
    upsert_cliente_mapping,
    validar_subcuenta_cliente,
    validar_subcuenta_ingreso,
)
from app.sources.base import Articulo, Cliente, DataSource, Filtros

# ─── Fuente de datos fake para los tests de sync ───────────────────────────


class _FakeSource(DataSource):
    def __init__(
        self,
        clientes: list[Cliente] | None = None,
        articulos: list[Articulo] | None = None,
    ) -> None:
        self._clientes = clientes or []
        self._articulos = articulos or []

    def get_facturas(self, filtros: Filtros):  # pragma: no cover - no usado aquí
        return []

    def get_cliente(self, codigo):  # pragma: no cover
        for c in self._clientes:
            if c.codigo == codigo:
                return c
        raise KeyError(codigo)

    def get_clientes(self):
        return list(self._clientes)

    def get_articulos(self):
        return list(self._articulos)


def _cliente(codigo='CLI00001', nombre='Cliente Uno', nif='12345678A'):
    return Cliente(codigo=codigo, nombre=nombre, nif=nif)


def _articulo(clave='CHIP', descripcion='Microchip canino'):
    return Articulo(clave=clave, descripcion=descripcion)


# ─── Validación de subcuentas ──────────────────────────────────────────────


def test_set_subcuenta_6_digitos_validos():
    assert validar_subcuenta_cliente('430001') == '430001'
    assert validar_subcuenta_cliente('430999') == '430999'
    assert validar_subcuenta_ingreso('700001') == '700001'
    assert validar_subcuenta_ingreso('705001') == '705001'
    assert validar_subcuenta_ingreso('755001') == '755001'


def test_set_subcuenta_7_digitos_error():
    with pytest.raises(SubcuentaInvalidaError):
        validar_subcuenta_cliente('4300001')
    with pytest.raises(SubcuentaInvalidaError):
        validar_subcuenta_ingreso('7000001')


def test_set_subcuenta_letras_error():
    with pytest.raises(SubcuentaInvalidaError):
        validar_subcuenta_cliente('430A01')
    with pytest.raises(SubcuentaInvalidaError):
        validar_subcuenta_ingreso('70A001')


def test_set_subcuenta_prefijo_invalido_error():
    with pytest.raises(SubcuentaInvalidaError):
        validar_subcuenta_cliente('440001')  # prefijo 44
    with pytest.raises(SubcuentaInvalidaError):
        validar_subcuenta_ingreso('600001')  # prefijo 60


# ─── CRUD y revisado ───────────────────────────────────────────────────────


def test_upsert_cliente_inserta_y_actualiza(db_conn):
    assert upsert_cliente_mapping(
        db_conn, codigo='CLI001', nombre='A', nif=None,
    ) == 'insertado'
    assert upsert_cliente_mapping(
        db_conn, codigo='CLI001', nombre='A bis', nif='X',
    ) == 'actualizado'
    fila = get_cliente_mapping(db_conn, 'CLI001')
    assert fila['nombre'] == 'A bis'
    assert fila['nif'] == 'X'


def test_upsert_cliente_valida_subcuenta(db_conn):
    with pytest.raises(SubcuentaInvalidaError):
        upsert_cliente_mapping(db_conn, codigo='CLI001', nombre='A', subcuenta_a3='999')


def test_set_subcuenta_cliente_persiste(db_conn):
    upsert_cliente_mapping(db_conn, codigo='CLI001', nombre='A')
    set_subcuenta_cliente(db_conn, 'CLI001', '430001')
    fila = get_cliente_mapping(db_conn, 'CLI001')
    assert fila['subcuenta_a3'] == '430001'


def test_set_subcuenta_cliente_no_existe(db_conn):
    with pytest.raises(KeyError):
        set_subcuenta_cliente(db_conn, 'NOPE', '430001')


def test_marcar_revisado(db_conn):
    upsert_cliente_mapping(db_conn, codigo='CLI001', nombre='A')
    marcar_cliente_revisado(db_conn, 'CLI001')
    fila = get_cliente_mapping(db_conn, 'CLI001')
    assert fila['revisado'] == 1
    assert fila['fecha_revision'] is not None


def test_marcar_revisado_no_existe(db_conn):
    with pytest.raises(KeyError):
        marcar_cliente_revisado(db_conn, 'NOPE')


def test_progreso_clientes(db_conn):
    upsert_cliente_mapping(db_conn, codigo='CLI001', nombre='A')
    upsert_cliente_mapping(db_conn, codigo='CLI002', nombre='B')
    marcar_cliente_revisado(db_conn, 'CLI001')
    assert progreso_clientes(db_conn) == (1, 2)


def test_list_clientes_filtros(db_conn):
    upsert_cliente_mapping(db_conn, codigo='CLI001', nombre='A')
    upsert_cliente_mapping(db_conn, codigo='CLI002', nombre='B')
    marcar_cliente_revisado(db_conn, 'CLI001')

    assert len(list_clientes_mappings(db_conn, 'todos')) == 2
    assert len(list_clientes_mappings(db_conn, 'pendientes')) == 1
    assert len(list_clientes_mappings(db_conn, 'revisados')) == 1

    with pytest.raises(ValueError):
        list_clientes_mappings(db_conn, 'inventado')


def test_marcar_articulo_revisado(db_conn):
    from app.mapping.store import upsert_articulo_mapping
    upsert_articulo_mapping(db_conn, clave='CHIP', descripcion='Microchip')
    set_cuenta_articulo(db_conn, 'CHIP', '700001')
    marcar_articulo_revisado(db_conn, 'CHIP')
    fila = get_articulo_mapping(db_conn, 'CHIP')
    assert fila['revisado'] == 1
    assert fila['cuenta_a3'] == '700001'


# ─── Init idempotente (sync sin sobreescribir revisados) ──────────────────


def test_init_idempotente(db_conn):
    """sync_clientes() dos veces no sobreescribe filas revisadas."""
    source = _FakeSource(clientes=[_cliente()])
    r1 = sync_mod.sync_clientes(db_conn, source)
    assert r1.insertados == 1

    # Operador revisa y asigna subcuenta
    set_subcuenta_cliente(db_conn, 'CLI00001', '430001')
    marcar_cliente_revisado(db_conn, 'CLI00001')

    # Segundo sync con nombre cambiado en la fuente → NO debe tocar
    source2 = _FakeSource(clientes=[_cliente(nombre='Nombre cambiado')])
    r2 = sync_mod.sync_clientes(db_conn, source2)
    assert r2.saltados_revisados == 1

    fila = get_cliente_mapping(db_conn, 'CLI00001')
    assert fila['nombre'] == 'Cliente Uno', "revisado=1 no se debe refrescar"
    assert fila['subcuenta_a3'] == '430001'
    assert fila['revisado'] == 1


def test_sync_refresca_no_revisados(db_conn):
    source1 = _FakeSource(clientes=[_cliente(nombre='Antes')])
    sync_mod.sync_clientes(db_conn, source1)

    source2 = _FakeSource(clientes=[_cliente(nombre='Después')])
    r = sync_mod.sync_clientes(db_conn, source2)
    assert r.actualizados == 1
    assert get_cliente_mapping(db_conn, 'CLI00001')['nombre'] == 'Después'


def test_sync_articulos(db_conn):
    source = _FakeSource(articulos=[
        _articulo('CHIP', 'Microchip canino'),
        _articulo('CERT', 'Certificado salud'),
    ])
    r = sync_mod.sync_articulos(db_conn, source)
    assert r.insertados == 2
    assert get_articulo_mapping(db_conn, 'CHIP')['descripcion'] == 'Microchip canino'


# ─── CSV import/export ────────────────────────────────────────────────────


def _escribir_csv_clientes(ruta: Path, filas: list[dict]) -> None:
    import csv as _csv
    columnas = csv_handler.CLIENTE_FIELDS
    with open(ruta, 'w', encoding=csv_handler.ENCODING, newline='') as fh:
        writer = _csv.DictWriter(fh, fieldnames=columnas)
        writer.writeheader()
        for f in filas:
            writer.writerow({c: f.get(c, '') for c in columnas})


def test_importar_csv_carga_masiva(db_conn, tmp_path):
    ruta = tmp_path / 'clientes.csv'
    _escribir_csv_clientes(ruta, [
        {'codigo_gesdai': 'CLI001', 'nombre': 'Alpha', 'subcuenta_a3': '430001'},
        {'codigo_gesdai': 'CLI002', 'nombre': 'Beta', 'subcuenta_a3': '430002'},
        {'codigo_gesdai': 'CLI003', 'nombre': 'Gamma'},  # sin subcuenta, válido
    ])

    resultado = csv_handler.importar_mappings_clientes(db_conn, ruta)
    assert resultado.insertados == 3
    assert resultado.errores == []
    assert get_cliente_mapping(db_conn, 'CLI002')['subcuenta_a3'] == '430002'


def test_importar_csv_error_subcuenta_no_bloquea_lote(db_conn, tmp_path):
    ruta = tmp_path / 'clientes.csv'
    _escribir_csv_clientes(ruta, [
        {'codigo_gesdai': 'CLI001', 'nombre': 'Alpha', 'subcuenta_a3': '430001'},
        {'codigo_gesdai': 'CLI002', 'nombre': 'Beta', 'subcuenta_a3': 'BAD'},
        {'codigo_gesdai': 'CLI003', 'nombre': 'Gamma', 'subcuenta_a3': '430003'},
    ])
    resultado = csv_handler.importar_mappings_clientes(db_conn, ruta)
    assert resultado.insertados == 2
    assert len(resultado.errores) == 1
    assert 'línea 3' in resultado.errores[0]


def test_importar_csv_no_sobreescribe_revisados(db_conn, tmp_path):
    upsert_cliente_mapping(db_conn, codigo='CLI001', nombre='Original', subcuenta_a3='430001')
    marcar_cliente_revisado(db_conn, 'CLI001')

    ruta = tmp_path / 'clientes.csv'
    _escribir_csv_clientes(ruta, [
        {'codigo_gesdai': 'CLI001', 'nombre': 'Otro nombre', 'subcuenta_a3': '430999'},
    ])
    resultado = csv_handler.importar_mappings_clientes(db_conn, ruta)
    assert resultado.saltados_revisados == 1

    fila = get_cliente_mapping(db_conn, 'CLI001')
    assert fila['nombre'] == 'Original'
    assert fila['subcuenta_a3'] == '430001'


def test_importar_csv_falta_columna(db_conn, tmp_path):
    ruta = tmp_path / 'mal.csv'
    ruta.write_text('codigo_gesdai,nombre\nCLI001,Alpha\n', encoding=csv_handler.ENCODING)
    with pytest.raises(ValueError, match='faltan columnas'):
        csv_handler.importar_mappings_clientes(db_conn, ruta)


def test_csv_round_trip_preserva_revisado(db_conn, tmp_path):
    """Regresión: el round-trip exportar→borrar→importar debe preservar
    el flag `revisado` de las filas marcadas, incluyendo la fecha."""
    upsert_cliente_mapping(db_conn, codigo='CLI001', nombre='Alpha', subcuenta_a3='430001')
    marcar_cliente_revisado(db_conn, 'CLI001')

    ruta = tmp_path / 'clientes.csv'
    csv_handler.exportar_mappings_clientes(db_conn, ruta)
    db_conn.execute("DELETE FROM mappings_clientes")
    db_conn.commit()
    csv_handler.importar_mappings_clientes(db_conn, ruta)

    fila = get_cliente_mapping(db_conn, 'CLI001')
    assert fila['revisado'] == 1
    assert fila['fecha_revision'] is not None


def test_csv_round_trip_articulo_preserva_es_texto_libre_y_revisado(db_conn, tmp_path):
    """Regresión: el flag es_texto_libre y revisado deben sobrevivir al round-trip."""
    from app.mapping.store import upsert_articulo_mapping
    upsert_articulo_mapping(
        db_conn, clave='VISITA', descripcion='Visita clínica',
        cuenta_a3='755001', es_texto_libre=True,
    )
    marcar_articulo_revisado(db_conn, 'VISITA')

    ruta = tmp_path / 'articulos.csv'
    csv_handler.exportar_mappings_articulos(db_conn, ruta)
    db_conn.execute("DELETE FROM mappings_articulos")
    db_conn.commit()
    csv_handler.importar_mappings_articulos(db_conn, ruta)

    fila = get_articulo_mapping(db_conn, 'VISITA')
    assert fila['es_texto_libre'] == 1
    assert fila['revisado'] == 1


def test_sync_articulos_no_pisa_es_texto_libre(db_conn):
    """Regresión: una re-sincronización desde GESDAI no debe resetear el
    flag `es_texto_libre` que el operador marcó en la UI. La fuente no
    sabe del flag y pasa el default None → COALESCE preserva el valor."""
    from app.mapping.store import upsert_articulo_mapping
    upsert_articulo_mapping(
        db_conn, clave='CERT', descripcion='Certificado',
        cuenta_a3='755001', es_texto_libre=True,
    )

    source = _FakeSource(articulos=[_articulo('CERT', 'Certificado')])
    sync_mod.sync_articulos(db_conn, source)

    fila = get_articulo_mapping(db_conn, 'CERT')
    assert fila['es_texto_libre'] == 1, (
        "sync no debe pisar es_texto_libre marcado a mano por el operador"
    )
    assert fila['cuenta_a3'] == '755001'


def test_exportar_csv_round_trip(db_conn, tmp_path):
    upsert_cliente_mapping(
        db_conn, codigo='CLI001', nombre='Alpha', nif='12345678A', subcuenta_a3='430001',
    )
    upsert_cliente_mapping(
        db_conn, codigo='CLI002', nombre='Beta, con coma', nif='87654321B', subcuenta_a3='430002',
    )

    ruta = tmp_path / 'export.csv'
    csv_handler.exportar_mappings_clientes(db_conn, ruta)
    assert ruta.exists()

    # Vaciar y re-importar en una BD limpia
    db_conn.execute("DELETE FROM mappings_clientes")
    db_conn.commit()
    assert list_clientes_mappings(db_conn) == []

    resultado = csv_handler.importar_mappings_clientes(db_conn, ruta)
    assert resultado.insertados == 2
    assert resultado.errores == []

    filas = list_clientes_mappings(db_conn)
    assert {f['codigo_gesdai'] for f in filas} == {'CLI001', 'CLI002'}
    assert any(f['nombre'] == 'Beta, con coma' for f in filas)


# ─── Excel import/export ──────────────────────────────────────────────────


def test_importar_excel_carga_masiva(db_conn, tmp_path):
    # Creamos un .xlsx de entrada con openpyxl directamente
    from openpyxl import Workbook
    ruta = tmp_path / 'clientes.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.append(list(csv_handler.CLIENTE_FIELDS))
    ws.append(['CLI001', 'Alpha', '12345678A', '430001', '0', ''])
    ws.append(['CLI002', 'Beta', '', '430002', '0', 'notas'])
    wb.save(ruta)

    resultado = excel_handler.importar_mappings_clientes(db_conn, ruta)
    assert resultado.insertados == 2
    assert resultado.errores == []
    assert get_cliente_mapping(db_conn, 'CLI002')['notas'] == 'notas'


def test_exportar_excel_round_trip(db_conn, tmp_path):
    upsert_cliente_mapping(
        db_conn, codigo='CLI001', nombre='Alpha', nif='12345678A', subcuenta_a3='430001',
    )
    upsert_cliente_mapping(
        db_conn, codigo='CLI002', nombre='Beta', subcuenta_a3='430002',
    )

    ruta = tmp_path / 'export.xlsx'
    excel_handler.exportar_mappings_clientes(db_conn, ruta)
    assert ruta.exists()

    db_conn.execute("DELETE FROM mappings_clientes")
    db_conn.commit()

    resultado = excel_handler.importar_mappings_clientes(db_conn, ruta)
    assert resultado.insertados == 2
    filas = list_clientes_mappings(db_conn)
    codigos = {f['codigo_gesdai'] for f in filas}
    assert codigos == {'CLI001', 'CLI002'}


# ─── CRUD artículos (cierra huecos del store) ─────────────────────────────


def test_upsert_articulo_insert_update_y_revisado(db_conn):
    from app.mapping.store import upsert_articulo_mapping
    assert upsert_articulo_mapping(
        db_conn, clave='CHIP', descripcion='Microchip', cuenta_a3='700001',
    ) == 'insertado'
    assert upsert_articulo_mapping(
        db_conn, clave='CHIP', descripcion='Microchip v2',
    ) == 'actualizado'
    assert get_articulo_mapping(db_conn, 'CHIP')['descripcion'] == 'Microchip v2'

    marcar_articulo_revisado(db_conn, 'CHIP')
    assert upsert_articulo_mapping(
        db_conn, clave='CHIP', descripcion='Intento de sobreescribir',
    ) == 'saltado_revisado'
    assert get_articulo_mapping(db_conn, 'CHIP')['descripcion'] == 'Microchip v2'


def test_upsert_articulo_valida_cuenta(db_conn):
    from app.mapping.store import upsert_articulo_mapping
    with pytest.raises(SubcuentaInvalidaError):
        upsert_articulo_mapping(db_conn, clave='CHIP', cuenta_a3='999999')


def test_set_cuenta_articulo_no_existe(db_conn):
    with pytest.raises(KeyError):
        set_cuenta_articulo(db_conn, 'NOPE', '700001')


def test_marcar_articulo_revisado_no_existe(db_conn):
    with pytest.raises(KeyError):
        marcar_articulo_revisado(db_conn, 'NOPE')


def test_marcar_articulo_desmarcar(db_conn):
    from app.mapping.store import upsert_articulo_mapping
    upsert_articulo_mapping(db_conn, clave='CHIP', descripcion='X')
    marcar_articulo_revisado(db_conn, 'CHIP', revisado=True)
    marcar_articulo_revisado(db_conn, 'CHIP', revisado=False)
    fila = get_articulo_mapping(db_conn, 'CHIP')
    assert fila['revisado'] == 0
    assert fila['fecha_revision'] is None


def test_list_articulos_filtros(db_conn):
    from app.mapping.store import (
        list_articulos_mappings,
        progreso_articulos,
        upsert_articulo_mapping,
    )
    upsert_articulo_mapping(db_conn, clave='CHIP')
    upsert_articulo_mapping(db_conn, clave='CERT')
    marcar_articulo_revisado(db_conn, 'CERT')

    assert len(list_articulos_mappings(db_conn, 'todos')) == 2
    assert len(list_articulos_mappings(db_conn, 'pendientes')) == 1
    assert len(list_articulos_mappings(db_conn, 'revisados')) == 1
    assert progreso_articulos(db_conn) == (1, 2)

    with pytest.raises(ValueError):
        list_articulos_mappings(db_conn, 'inventado')


# ─── CSV artículos y keywords ──────────────────────────────────────────────


def test_importar_csv_articulos_round_trip(db_conn, tmp_path):
    import csv as _csv

    ruta = tmp_path / 'articulos.csv'
    with open(ruta, 'w', encoding=csv_handler.ENCODING, newline='') as fh:
        writer = _csv.DictWriter(fh, fieldnames=csv_handler.ARTICULO_FIELDS)
        writer.writeheader()
        writer.writerow({
            'clave_gesdai': 'CHIP', 'descripcion': 'Microchip',
            'cuenta_a3': '700001', 'es_texto_libre': '0', 'revisado': '0', 'notas': '',
        })
        writer.writerow({
            'clave_gesdai': 'VISITA', 'descripcion': 'Visita clínica',
            'cuenta_a3': '755001', 'es_texto_libre': '1', 'revisado': '0', 'notas': '',
        })

    resultado = csv_handler.importar_mappings_articulos(db_conn, ruta)
    assert resultado.insertados == 2
    assert get_articulo_mapping(db_conn, 'VISITA')['es_texto_libre'] == 1

    # Round-trip: exportar, vaciar, re-importar
    ruta_export = tmp_path / 'articulos_export.csv'
    csv_handler.exportar_mappings_articulos(db_conn, ruta_export)
    db_conn.execute("DELETE FROM mappings_articulos")
    db_conn.commit()
    resultado = csv_handler.importar_mappings_articulos(db_conn, ruta_export)
    assert resultado.insertados == 2


def test_importar_csv_articulos_error_cuenta(db_conn, tmp_path):
    import csv as _csv

    ruta = tmp_path / 'articulos.csv'
    with open(ruta, 'w', encoding=csv_handler.ENCODING, newline='') as fh:
        writer = _csv.DictWriter(fh, fieldnames=csv_handler.ARTICULO_FIELDS)
        writer.writeheader()
        writer.writerow({
            'clave_gesdai': '', 'descripcion': 'X',
            'cuenta_a3': '', 'es_texto_libre': '0', 'revisado': '0', 'notas': '',
        })
        writer.writerow({
            'clave_gesdai': 'CHIP', 'descripcion': 'Chip',
            'cuenta_a3': '999999', 'es_texto_libre': '0', 'revisado': '0', 'notas': '',
        })

    resultado = csv_handler.importar_mappings_articulos(db_conn, ruta)
    assert resultado.insertados == 0
    assert len(resultado.errores) == 2


def test_csv_keywords_import_export_round_trip(db_conn, tmp_path):
    from app.mapping.keywords import add_keyword, load_keywords

    add_keyword(db_conn, keyword='chip', cuenta_a3='700001', prioridad=30)
    add_keyword(db_conn, keyword='cert', cuenta_a3='755001', prioridad=50)

    ruta = tmp_path / 'keywords.csv'
    csv_handler.exportar_keywords(db_conn, ruta)
    assert ruta.exists()

    db_conn.execute("DELETE FROM keywords_articulos")
    db_conn.commit()
    assert load_keywords(db_conn) == []

    resultado = csv_handler.importar_keywords(db_conn, ruta)
    assert resultado.insertados == 2
    assert resultado.errores == []

    nuevas = load_keywords(db_conn)
    assert [k.keyword for k in nuevas] == ['cert', 'chip']


def test_csv_keywords_import_error_no_bloquea(db_conn, tmp_path):
    ruta = tmp_path / 'keywords.csv'
    ruta.write_text(
        'keyword,cuenta_a3,prioridad,activo\n'
        'chip,700001,10,1\n'
        ',755001,10,1\n'        # keyword vacía
        'cert,XXX,10,1\n',       # cuenta inválida
        encoding=csv_handler.ENCODING,
    )
    resultado = csv_handler.importar_keywords(db_conn, ruta)
    assert resultado.insertados == 1
    assert len(resultado.errores) == 2


def test_csv_exportar_historial(db_conn, tmp_path):
    db_conn.execute(
        """INSERT INTO exportaciones
           (fecha_export, fecha_desde, fecha_hasta, num_facturas,
            dat_fichero, dat_hash, app_version)
           VALUES (?, ?, ?, ?, ?, ?, ?)""",
        ('2026-01-31', '2026-01-01', '2026-01-31', 42, '/ex.dat', 'abc', '0.1.0'),
    )
    ruta = tmp_path / 'historial.csv'
    csv_handler.exportar_historial(db_conn, ruta)
    contenido = ruta.read_text(encoding=csv_handler.ENCODING)
    assert 'fecha_export' in contenido
    assert '42' in contenido


# ─── Excel artículos + errores ─────────────────────────────────────────────


def test_excel_articulos_round_trip(db_conn, tmp_path):
    from app.mapping.store import upsert_articulo_mapping

    upsert_articulo_mapping(db_conn, clave='CHIP', descripcion='Microchip', cuenta_a3='700001')
    upsert_articulo_mapping(db_conn, clave='CERT', descripcion='Certificado', cuenta_a3='755001')

    ruta = tmp_path / 'articulos.xlsx'
    excel_handler.exportar_mappings_articulos(db_conn, ruta)
    assert ruta.exists()

    db_conn.execute("DELETE FROM mappings_articulos")
    db_conn.commit()

    resultado = excel_handler.importar_mappings_articulos(db_conn, ruta)
    assert resultado.insertados == 2
    assert get_articulo_mapping(db_conn, 'CHIP')['cuenta_a3'] == '700001'


def test_excel_importar_falta_columna(db_conn, tmp_path):
    from openpyxl import Workbook
    ruta = tmp_path / 'mal.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.append(['codigo_gesdai', 'nombre'])  # faltan columnas
    ws.append(['CLI001', 'Alpha'])
    wb.save(ruta)

    with pytest.raises(ValueError, match='faltan columnas'):
        excel_handler.importar_mappings_clientes(db_conn, ruta)


def test_excel_importar_hoja_vacia(db_conn, tmp_path):
    from openpyxl import Workbook
    ruta = tmp_path / 'vacio.xlsx'
    wb = Workbook()
    wb.active.delete_rows(1, 10)  # sin ninguna fila
    wb.save(ruta)
    with pytest.raises(ValueError, match='vacía'):
        excel_handler.importar_mappings_clientes(db_conn, ruta)


def test_excel_importar_salta_filas_en_blanco(db_conn, tmp_path):
    from openpyxl import Workbook

    ruta = tmp_path / 'clientes.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.append(list(csv_handler.CLIENTE_FIELDS))
    ws.append(['CLI001', 'Alpha', '', '430001', '0', ''])
    ws.append([None, None, None, None, None, None])  # fila en blanco
    ws.append(['CLI002', 'Beta', '', '430002', '0', ''])
    wb.save(ruta)

    resultado = excel_handler.importar_mappings_clientes(db_conn, ruta)
    assert resultado.insertados == 2
    assert resultado.errores == []


def test_excel_exportar_historial(db_conn, tmp_path):
    db_conn.execute(
        """INSERT INTO exportaciones
           (fecha_export, fecha_desde, fecha_hasta, num_facturas,
            dat_fichero, dat_hash, app_version)
           VALUES (?, ?, ?, ?, ?, ?, ?)""",
        ('2026-01-31', '2026-01-01', '2026-01-31', 10, '/ex.dat', 'h', '0.1.0'),
    )
    ruta = tmp_path / 'historial.xlsx'
    excel_handler.exportar_historial(db_conn, ruta)
    assert ruta.exists() and ruta.stat().st_size > 0
